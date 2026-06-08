# --- BEGIN GENERATED FILE METADATA ---
# Authors: Ansh Raj
# Last edited: 2026-06-05T10:54:18Z
# --- END GENERATED FILE METADATA ---

"""Excel streaming reader using openpyxl read-only mode."""

from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from category1.models.schemas import ColumnSchema, ConnectionConfig, DatasetSchema
from category1.readers.base import StreamingReader


class ExcelReader(StreamingReader):
    def __init__(self, config: ConnectionConfig):
        super().__init__(config)
        self.file_path = Path(config.file_path or "")
        opts = config.file_options
        self.sheet_name = opts.get("sheet_name", None)
        self.has_header = opts.get("has_header", True)

    def get_schema(self) -> DatasetSchema:
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active
        columns: list[ColumnSchema] = []
        if self.has_header:
            for i, cell in enumerate(next(ws.iter_rows(max_row=1, values_only=True))):
                name = str(cell) if cell is not None else f"col_{i}"
                columns.append(ColumnSchema(name=name, data_type="string", position=i))
        wb.close()
        return DatasetSchema(columns=columns)

    def read_chunks(self, chunk_size: int) -> Iterator[list[dict[str, Any]]]:
        schema = self.get_schema()
        col_names = [c.name for c in schema.columns]
        chunk: list[dict[str, Any]] = []

        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active
        start_row = 2 if self.has_header else 1

        for row in ws.iter_rows(min_row=start_row, values_only=True):
            record = {}
            for i, name in enumerate(col_names):
                record[name] = row[i] if i < len(row) else None
            chunk.append(record)
            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []
        wb.close()
        if chunk:
            yield chunk
